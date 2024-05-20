import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.stats import norm
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import load_workbook

# Background image setup
page_bg_img = """
<style>
[data-testid="stAppViewContainer"] {
background: url("https://i.imgur.com/kox6xPx.png");
background-size: cover;
}
</style>
"""
st.markdown(page_bg_img, unsafe_allow_html=True)

if 'button_clicked' not in st.session_state:
    st.session_state.button_clicked = False

if st.button('Press Me', key='press_me_button', on_click=lambda: st.session_state.update(button_clicked=True)):
    st.session_state.button_clicked = True

if st.session_state.button_clicked:
    compare_policies = st.checkbox('Compare Policies')

    def generate_demand(distribution, duration, mean, std_dev):
        if distribution == "Normal":
            return np.maximum(np.random.normal(loc=mean, scale=std_dev, size=duration).round(0).astype(int), 0)
        elif distribution == "Poisson":
            return np.random.poisson(lam=mean, size=duration)
        elif distribution == "Uniform":
            return np.random.uniform(low=mean - std_dev, high=mean + std_dev, size=duration).astype(int)

    def simulate_inventory_RS(duration, demand, mean_demand, std_dev, lead_time, review_period, service_level):
        z = norm.ppf(service_level)
        x_std = std_dev * np.sqrt(lead_time + review_period)
        Ss = np.round(x_std * z).astype(int)
        Q = mean_demand * review_period
        Cs = Q / 2
        Is = mean_demand * lead_time
        S = Ss + Q + Is

        hand = np.zeros(duration, dtype=int)
        transit = np.zeros((duration, lead_time + 1), dtype=int)
        shortages = np.zeros(duration, dtype=int)

        hand[0] = S - demand[0]
        transit[1, -1] = demand[0]

        stock_out_period = np.full(duration, False, dtype=bool)
        stock_out_cycle = []

        for t in range(1, duration):
            if transit[t-1, 0] > 0:
                stock_out_cycle.append(stock_out_period[t-1])
            hand[t] = hand[t-1] - demand[t] + transit[t-1, 0]
            shortages[t] = max(0, demand[t] - hand[t])
            stock_out_period[t] = hand[t] < 0
            transit[t, :-1] = transit[t-1, 1:]
            if 0 == t % review_period:
                net = hand[t] + transit[t].sum()
                transit[t, lead_time] = S - net

        SL_alpha = (1 - sum(stock_out_cycle) / len(stock_out_cycle)) * 100
        SL_period = (1 - sum(stock_out_period) / duration) * 100

        return hand.astype(int), transit.astype(int), shortages.astype(int), stock_out_period, SL_alpha, SL_period, Ss, Cs, Is

    def simulate_inventory_sQ(duration, demand, mean_demand, std_dev, lead_time, service_level, s, Q):
        z = norm.ppf(service_level)
        hand = np.zeros(duration, dtype=int)
        transit = np.zeros((duration, lead_time + 1), dtype=int)
        shortages = np.zeros(duration, dtype=int)

        hand[0] = s + Q - demand[0]
        transit[1, -1] = Q

        stock_out_period = np.full(duration, False, dtype=bool)
        stock_out_cycle = []

        for t in range(1, duration):
            if transit[t-1, 0] > 0:
                stock_out_cycle.append(stock_out_period[t-1])
            hand[t] = hand[t-1] - demand[t] + transit[t-1, 0]
            shortages[t] = max(0, demand[t] - hand[t])
            stock_out_period[t] = hand[t] < 0
            transit[t, :-1] = transit[t-1, 1:]
            if hand[t] < s:
                transit[t, lead_time] = Q

        SL_alpha = (1 - sum(stock_out_cycle) / len(stock_out_cycle)) * 100
        SL_period = (1 - sum(stock_out_period) / duration) * 100

        return hand.astype(int), transit.astype(int), shortages.astype(int), stock_out_period, SL_alpha, SL_period

    if compare_policies:
        st.title("Inventory Management - Compare Policies")

        col1, col2 = st.columns(2)

        with col1:
            st.header("Policy 1")
            duration1 = st.number_input("Duration (days)", value=30, key="duration1")
            mean_demand1 = st.number_input("Demand Mean:", value=50, key="mean_demand1")
            std_dev1 = st.number_input("Demand Std Dev:", value=10, key="std_dev1")
            lead_time1 = st.number_input("Lead Time (days):", value=5, key="lead_time1")
            policy1 = st.selectbox("Policy:", ["R,S", "s,Q"], key="policy1")
            distribution1 = st.selectbox("Demand Distribution:", ["Normal", "Poisson", "Uniform"], key="distribution1")

            if policy1 == "R,S":
                review_period1 = st.number_input("Review Period (R):", value=10, key="R1")
            elif policy1 == "s,Q":
                s1 = st.number_input("Reorder Point (s):", value=20, key="s1")
                Q1 = st.number_input("Order Quantity (Q):", value=40, key="Q1")

        with col2:
            st.header("Policy 2")
            duration2 = st.number_input("Duration (days)", value=30, key="duration2")
            mean_demand2 = st.number_input("Demand Mean:", value=50, key="mean_demand2")
            std_dev2 = st.number_input("Demand Std Dev:", value=10, key="std_dev2")
            lead_time2 = st.number_input("Lead Time (days):", value=5, key="lead_time2")
            policy2 = st.selectbox("Policy:", ["R,S", "s,Q"], key="policy2")
            distribution2 = st.selectbox("Demand Distribution:", ["Normal", "Poisson", "Uniform"], key="distribution2")

            if policy2 == "R,S":
                review_period2 = st.number_input("Review Period (R):", value=10, key="R2")
            elif policy2 == "s,Q":
                s2 = st.number_input("Reorder Point (s):", value=20, key="s2")
                Q2 = st.number_input("Order Quantity (Q):", value=40, key="Q2")

        service_level = st.slider('Service Level:', 0.80, 1.00, 0.95)

        if st.button("Run Simulation"):
            demand_data = generate_demand(distribution1, max(duration1, duration2), mean_demand1, std_dev1)

            if policy1 == "R,S":
                inventory_levels1, in_transit1, shortages1, stock_out_period1, SL_alpha1, SL_period1, Ss1, Cs1, Is1 = simulate_inventory_RS(
                    duration1, demand_data[:duration1], mean_demand1, std_dev1, lead_time1, review_period1, service_level)
            elif policy1 == "s,Q":
                inventory_levels1, in_transit1, shortages1, stock_out_period1, SL_alpha1, SL_period1 = simulate_inventory_sQ(
                    duration1, demand_data[:duration1], mean_demand1, std_dev1, lead_time1, service_level, s1, Q1)

            if policy2 == "R,S":
                inventory_levels2, in_transit2, shortages2, stock_out_period2, SL_alpha2, SL_period2, Ss2, Cs2, Is2 = simulate_inventory_RS(
                    duration2, demand_data[:duration2], mean_demand2, std_dev2, lead_time2, review_period2, service_level)
            elif policy2 == "s,Q":
                inventory_levels2, in_transit2, shortages2, stock_out_period2, SL_alpha2, SL_period2 = simulate_inventory_sQ(
                    duration2, demand_data[:duration2], mean_demand2, std_dev2, lead_time2, service_level, s2, Q2)

            # Plotting results
            fig, ax = plt.subplots()
            ax.plot(inventory_levels1, label=f'Inventory Level (Policy 1: {policy1})')
            ax.plot(inventory_levels2, label=f'Inventory Level (Policy 2: {policy2})')

            if policy1 == "R,S":
                ax.axhline(y=Ss1, color='r', linestyle='--', label='R,S Policy Reorder Point')
            if policy2 == "R,S":
                ax.axhline(y=Ss2, color='r', linestyle='--', label='R,S Policy Reorder Point')
            if policy1 == "s,Q":
                ax.axhline(y=s1, color='b', linestyle='--', label='s,Q Policy Reorder Point')
            if policy2 == "s,Q":
                ax.axhline(y=s2, color='b', linestyle='--', label='s,Q Policy Reorder Point')

            ax.set_xlabel('Time (days)')
            ax.set_ylabel('Inventory Level')
            ax.legend()
            st.pyplot(fig)

            # Calculate and display stocks
            st.subheader('Stock Calculations')
            if policy1 == 's,Q':
                cycle_stock1 = Q1 / 2
                pipeline_stock1 = mean_demand1 * lead_time1
                safety_stock1 = std_dev1 * np.sqrt(lead_time1) * norm.ppf(service_level)
            else:  # R,S
                cycle_stock1 = mean_demand1 * review_period1 / 2
                pipeline_stock1 = mean_demand1 * lead_time1
                safety_stock1 = std_dev1 * np.sqrt(lead_time1 + review_period1) * norm.ppf(service_level)

            st.write(f'Policy 1 ({policy1})')
            st.write('Cycle Stock:', cycle_stock1)
            st.write('Pipeline Stock:', pipeline_stock1)
            st.write('Safety Stock:', safety_stock1)

            if policy2 == 's,Q':
                cycle_stock2 = Q2 / 2
                pipeline_stock2 = mean_demand2 * lead_time2
                safety_stock2 = std_dev2 * np.sqrt(lead_time2) * norm.ppf(service_level)
            else:  # R,S
                cycle_stock2 = mean_demand2 * review_period2 / 2
                pipeline_stock2 = mean_demand2 * lead_time2
                safety_stock2 = std_dev2 * np.sqrt(lead_time2 + review_period2) * norm.ppf(service_level)

            st.write(f'Policy 2 ({policy2})')
            st.write('Cycle Stock:', cycle_stock2)
            st.write('Pipeline Stock:', pipeline_stock2)
            st.write('Safety Stock:', safety_stock2)

            # Writing results to Excel
            results_df1 = pd.DataFrame({
                'Time': range(duration1),
                'Demand': demand_data[:duration1],
                'On-hand': inventory_levels1,
                'In-transit': list(in_transit1),
                'Shortages': shortages1
            })

            results_df2 = pd.DataFrame({
                'Time': range(duration2),
                'Demand': demand_data[:duration2],
                'On-hand': inventory_levels2,
                'In-transit': list(in_transit2),
                'Shortages': shortages2
            })

            # Ensure sheet names are valid by removing any special characters
            valid_policy1 = ''.join(e for e in policy1 if e.isalnum())
            valid_policy2 = ''.join(e for e in policy2 if e.isalnum())

            file_path = 'inventorycontrol_comparison.xlsx'
            with pd.ExcelWriter(file_path) as writer:
                results_df1.to_excel(writer, sheet_name=f'Policy1_{valid_policy1}', index=False)
                results_df2.to_excel(writer, sheet_name=f'Policy2_{valid_policy2}', index=False)

            # Load the workbook to apply formatting
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Apply formatting
                header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply alternating row colors
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        if cell.row % 2 == 0:
                            cell.fill = PatternFill(start_color='FFEBEB', end_color='FFEBEB', fill_type='solid')

            wb.save(file_path)

            st.success(f"Results saved to {file_path}")
            st.write(f"Cycle Service Level for Policy 1: {SL_alpha1:.2f}%")
            st.write(f"Period Service Level for Policy 1: {SL_period1:.2f}%")
            st.write(f"Cycle Service Level for Policy 2: {SL_alpha2:.2f}%")
            st.write(f"Period Service Level for Policy 2: {SL_period2:.2f}%")
            st.download_button('Download Comparison Report', data=open(file_path, 'rb').read(), file_name=file_path, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    else:
        st.title("Inventory Management - Single Policy")

        st.header("Policy")
        duration = st.number_input("Duration (days)", value=30, key="duration")
        mean_demand = st.number_input("Demand Mean:", value=50, key="mean_demand")
        std_dev = st.number_input("Demand Std Dev:", value=10, key="std_dev")
        lead_time = st.number_input("Lead Time (days):", value=5, key="lead_time")
        policy = st.selectbox("Policy:", ["R,S", "s,Q"], key="policy")
        distribution = st.selectbox("Demand Distribution:", ["Normal", "Poisson", "Uniform"], key="distribution")

        if policy == "R,S":
            review_period = st.number_input("Review Period (R):", value=10, key="R")
        elif policy == "s,Q":
            s = st.number_input("Reorder Point (s):", value=20, key="s")
            Q = st.number_input("Order Quantity (Q):", value=40, key="Q")

        service_level = st.slider('Service Level:', 0.80, 1.00, 0.95)

        if st.button("Run Simulation"):
            demand_data = generate_demand(distribution, duration, mean_demand, std_dev)

            if policy == "R,S":
                inventory_levels, in_transit, shortages, stock_out_period, SL_alpha, SL_period, Ss, Cs, Is = simulate_inventory_RS(
                    duration, demand_data, mean_demand, std_dev, lead_time, review_period, service_level)
            elif policy == "s,Q":
                inventory_levels, in_transit, shortages, stock_out_period, SL_alpha, SL_period = simulate_inventory_sQ(
                    duration, demand_data, mean_demand, std_dev, lead_time, service_level, s, Q)

            # Plotting results
            fig, ax = plt.subplots()
            ax.plot(inventory_levels, label=f'Inventory Level (Policy: {policy})')

            if policy == "R,S":
                ax.axhline(y=Ss, color='r', linestyle='--', label='R,S Policy Reorder Point')
            if policy == "s,Q":
                ax.axhline(y=s, color='b', linestyle='--', label='s,Q Policy Reorder Point')

            ax.set_xlabel('Time (days)')
            ax.set_ylabel('Inventory Level')
            ax.legend()
            st.pyplot(fig)

            # Calculate and display stocks
            st.subheader('Stock Calculations')
            if policy == 's,Q':
                cycle_stock = Q / 2
                pipeline_stock = mean_demand * lead_time
                safety_stock = std_dev * np.sqrt(lead_time) * norm.ppf(service_level)
            else:  # R,S
                cycle_stock = mean_demand * review_period / 2
                pipeline_stock = mean_demand * lead_time
                safety_stock = std_dev * np.sqrt(lead_time + review_period) * norm.ppf(service_level)

            st.write(f'Policy ({policy})')
            st.write('Cycle Stock:', cycle_stock)
            st.write('Pipeline Stock:', pipeline_stock)
            st.write('Safety Stock:', safety_stock)

            # Writing results to Excel
            results_df = pd.DataFrame({
                'Time': range(duration),
                'Demand': demand_data,
                'On-hand': inventory_levels,
                'In-transit': list(in_transit),
                'Shortages': shortages
            })

            # Ensure sheet names are valid by removing any special characters
            valid_policy = ''.join(e for e in policy if e.isalnum())

            file_path = 'inventorycontrol_singlepolicy.xlsx'
            with pd.ExcelWriter(file_path) as writer:
                results_df.to_excel(writer, sheet_name=f'Policy_{valid_policy}', index=False)

            # Load the workbook to apply formatting
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Apply formatting
                header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply alternating row colors
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        if cell.row % 2 == 0:
                            cell.fill = PatternFill(start_color='FFEBEB', end_color='FFEBEB', fill_type='solid')

            wb.save(file_path)

            st.success(f"Results saved to {file_path}")
            st.write(f"Cycle Service Level: {SL_alpha:.2f}%")
            st.write(f"Period Service Level: {SL_period:.2f}%")
            st.download_button('Download Report', data=open(file_path, 'rb').read(), file_name=file_path, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            drive_link = "https://drive.google.com/file/d/1qxn0vv2nBxTgH_jAliVltXQ_XU6LUQfp/view?usp=sharing"
            if st.button('User Manual'):
                st.markdown(f"[User Manual]({drive_link})", unsafe_allow_html=True)

